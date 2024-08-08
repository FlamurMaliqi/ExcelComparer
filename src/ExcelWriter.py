from openpyxl.styles import PatternFill
from ExcelFile import ExcelFile
import tkinter
from tkinter import filedialog


# Die verglichenen Excel-Dateien müssen in einer Datei zusammengefasst gespeichert werden.
# Dabei sollen entsprechende Merkmale wie entfernte, eingefügte oder verschobene Object-Ids markiert werden.
class ExcelWriter:
    def __init__(self, old: ExcelFile, check: ExcelFile):
        self.old = old
        self.check = check

    # Färbt die Headingzeilen in einer gegebenen Farbe ein.
    # Dient zur Erkennung wo sich evtl. neue oder verschobene Headings befinden
    def heading_colouring(self, file, colour):
        for i in range(file.headings):
            heading_color = PatternFill(fgColor=colour, fill_type='solid')
            for col in range(max(self.old.max_column, self.check.max_row)):
                if col > 0:
                    self.check.worksheet.cell(column=col, row=file.heading_line_rows[i][1]).fill = heading_color

    def list_colouring(self, lst: list, colour: str):
        filler = PatternFill(fgColor=colour, fill_type='solid')
        for i in lst:
            for col in range(1, max(self.old.max_column, self.check.max_column)):
                self.check.worksheet.cell(column=col, row=i[0]).fill = filler

    def add_test_case_id(self, lst: list, col_val: int, value: str):
        cnt = 0
        for entry in lst:
            cnt += 1
            self.check.worksheet.cell(column=col_val, row=entry).value = (value + "_" + str(cnt))

    def add_new_row(self, lst: list, col_val: int, colour: str):
        cnt = 0
        filler = PatternFill(fgColor=colour, fill_type='solid')
        for entry in lst:
            cnt += 1
            self.check.worksheet.insert_rows(entry[0] + cnt)
            self.check.worksheet.cell(column=col_val, row=(entry[0] + cnt)).value = entry[1]
            self.check.worksheet.cell(column=col_val, row=(entry[0] + cnt)).fill = filler

    def delete_column(self, id):
        if id != 0:
            self.check.worksheet.delete_cols(id)

    # Funktion erstellt neue excel-Datei und kopiert dabei Inhalt von check.xlsx in result.xlsx.
    # Dabei wird der bearbeitete Inhalt von Check nie gespeichert und somit im Endeffekt nie verändert.
    # Filedialog ruft den Nutzer zum Speichern auf und lässt ihn den Ort dafür festlegen.
    def create_result(self, text):
        root = tkinter.Tk()
        root.withdraw()
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=(("Text files", "*.xlsx"), ("All files", "*.*")),
                                                 title="Save Result",
                                                 initialfile="processed_" + text)
        if file_path:
            # Create a new workbook
            self.check.workbook.save(file_path)
            print("Excel file saved at:", file_path)
        else:
            print("No file saved")
